"""Analytics tracking sheet generator for performance measurement"""

import csv
from pathlib import Path
from typing import List

from ..models.post import Post
from ..models.posting_schedule import PostingSchedule
from ..utils.logger import logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class AnalyticsTracker:
    """Generates analytics tracking spreadsheets for post performance"""

    # Column definitions
    COLUMNS = [
        "Post #",
        "Template",
        "Platform",
        "Scheduled Date",
        "Scheduled Time",
        "Word Count",
        "Has CTA",
        "Keywords",
        "Post URL",
        "Impressions",
        "Engagement",
        "Clicks",
        "Shares",
        "Comments",
        "Leads Generated",
        "Notes",
    ]

    def __init__(self):
        """Initialize analytics tracker"""

    def create_tracking_sheet(
        self,
        posts: List[Post],
        schedule: PostingSchedule,
        output_path: Path,
        format: str = "csv",
    ) -> Path:
        """
        Generate performance tracking spreadsheet

        Args:
            posts: List of Post objects
            schedule: PostingSchedule with dates/times
            output_path: Path for output file
            format: "csv" or "xlsx"

        Returns:
            Path to generated file
        """
        logger.info(f"Generating {format.upper()} analytics tracker for {len(posts)} posts")

        if format == "csv":
            return self._create_csv(posts, schedule, output_path)
        elif format == "xlsx":
            if not OPENPYXL_AVAILABLE:
                logger.warning(
                    "openpyxl not installed - falling back to CSV. "
                    "Install with: pip install openpyxl"
                )
                # Change extension to csv
                csv_path = output_path.with_suffix(".csv")
                return self._create_csv(posts, schedule, csv_path)
            return self._create_xlsx(posts, schedule, output_path)
        else:
            raise ValueError(f"Unsupported format: {format}. Use 'csv' or 'xlsx'")

    def _create_csv(self, posts: List[Post], schedule: PostingSchedule, output_path: Path) -> Path:
        """Create CSV tracking sheet"""
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Header row
            writer.writerow(self.COLUMNS)

            # Data rows
            for i, post in enumerate(posts):
                # Find corresponding scheduled post
                scheduled_post = next(
                    (sp for sp in schedule.scheduled_posts if sp.post_id == i + 1),
                    None,
                )

                # Prepare row data
                row = [
                    i + 1,  # Post #
                    post.template_name,  # Template
                    (
                        scheduled_post.platform if scheduled_post else post.target_platform or ""
                    ),  # Platform
                    (
                        scheduled_post.scheduled_date.strftime("%Y-%m-%d") if scheduled_post else ""
                    ),  # Scheduled Date
                    (
                        scheduled_post.scheduled_time.strftime("%H:%M")
                        if scheduled_post and scheduled_post.scheduled_time
                        else ""
                    ),  # Scheduled Time
                    post.word_count,  # Word Count
                    "Yes" if post.has_cta else "No",  # Has CTA
                    ", ".join(getattr(post, "keywords_used", [])[:3])
                    if getattr(post, "keywords_used", [])
                    else "",  # Keywords
                    "",  # Post URL (blank)
                    "",  # Impressions (blank)
                    "",  # Engagement (blank)
                    "",  # Clicks (blank)
                    "",  # Shares (blank)
                    "",  # Comments (blank)
                    "",  # Leads Generated (blank)
                    "",  # Notes (blank)
                ]

                writer.writerow(row)

        logger.info(f"CSV tracker saved to {output_path}")
        return output_path

    def _create_xlsx(self, posts: List[Post], schedule: PostingSchedule, output_path: Path) -> Path:
        """Create Excel tracking sheet with formulas and formatting"""
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()

        # Create data sheet
        ws_data = wb.active
        ws_data.title = "Post Tracker"

        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")

        # === DATA SHEET ===

        # Header row styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Write headers
        for col_num, header in enumerate(self.COLUMNS, 1):
            cell = ws_data.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for i, post in enumerate(posts, 2):  # Start at row 2 (row 1 is header)
            # Find corresponding scheduled post
            scheduled_post = next(
                (sp for sp in schedule.scheduled_posts if sp.post_id == i - 1),
                None,
            )

            # Column A: Post #
            ws_data[f"A{i}"] = i - 1

            # Column B: Template
            ws_data[f"B{i}"] = post.template_name

            # Column C: Platform
            ws_data[f"C{i}"] = (
                scheduled_post.platform if scheduled_post else post.target_platform or ""
            )

            # Column D: Scheduled Date
            if scheduled_post:
                ws_data[f"D{i}"] = scheduled_post.scheduled_date.strftime("%Y-%m-%d")

            # Column E: Scheduled Time
            if scheduled_post and scheduled_post.scheduled_time:
                ws_data[f"E{i}"] = scheduled_post.scheduled_time.strftime("%H:%M")

            # Column F: Word Count
            ws_data[f"F{i}"] = post.word_count

            # Column G: Has CTA
            ws_data[f"G{i}"] = "Yes" if post.has_cta else "No"

            # Column H: Keywords
            keywords = getattr(post, "keywords_used", [])
            ws_data[f"H{i}"] = ", ".join(keywords[:3]) if keywords else ""

            # Columns I-P: Leave blank for manual entry

        # Add formulas for engagement rate (if data is entered)
        # Formula: Engagement Rate = (Clicks + Shares + Comments) / Impressions
        # We'll add this as a note in the header

        # Auto-adjust column widths
        for col in range(1, len(self.COLUMNS) + 1):
            ws_data.column_dimensions[get_column_letter(col)].width = 15

        # Make specific columns wider
        ws_data.column_dimensions["B"].width = 25  # Template
        ws_data.column_dimensions["H"].width = 30  # Keywords
        ws_data.column_dimensions["I"].width = 40  # Post URL
        ws_data.column_dimensions["P"].width = 40  # Notes

        # === SUMMARY SHEET ===

        # Title
        ws_summary["A1"] = "Performance Summary"
        ws_summary["A1"].font = Font(bold=True, size=14)

        ws_summary["A3"] = "Total Posts:"
        ws_summary["B3"] = f"=COUNTA('Post Tracker'!A2:A{len(posts)+1})"

        ws_summary["A4"] = "Posts Published:"
        ws_summary[
            "B4"
        ] = f"=COUNTIF('Post Tracker'!I2:I{len(posts)+1},\"<>\")"  # Count non-empty URLs

        ws_summary["A5"] = "Total Impressions:"
        ws_summary["B5"] = f"=SUM('Post Tracker'!J2:J{len(posts)+1})"

        ws_summary["A6"] = "Total Engagement:"
        ws_summary["B6"] = f"=SUM('Post Tracker'!K2:K{len(posts)+1})"

        ws_summary["A7"] = "Total Clicks:"
        ws_summary["B7"] = f"=SUM('Post Tracker'!L2:L{len(posts)+1})"

        ws_summary["A8"] = "Total Shares:"
        ws_summary["B8"] = f"=SUM('Post Tracker'!M2:M{len(posts)+1})"

        ws_summary["A9"] = "Total Comments:"
        ws_summary["B9"] = f"=SUM('Post Tracker'!N2:N{len(posts)+1})"

        ws_summary["A10"] = "Total Leads:"
        ws_summary["B10"] = f"=SUM('Post Tracker'!O2:O{len(posts)+1})"

        ws_summary["A12"] = "Average Engagement Rate:"
        ws_summary["B12"] = "=IF(B5>0,B6/B5,0)"
        ws_summary["B12"].number_format = "0.00%"

        # Bold labels
        for row in range(3, 13):
            ws_summary[f"A{row}"].font = Font(bold=True)

        # Auto-adjust column widths
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 15

        # Instructions
        ws_summary["A15"] = "Instructions:"
        ws_summary["A15"].font = Font(bold=True, size=12)

        instructions = [
            "1. After posting, fill in Post URL in column I",
            "2. Track performance metrics (columns J-O) weekly",
            "3. Add notes in column P for insights or learnings",
            "4. Summary metrics auto-calculate as you enter data",
            "5. Sort by Impressions or Engagement to find top performers",
        ]

        for i, instruction in enumerate(instructions, 16):
            ws_summary[f"A{i}"] = instruction

        # Save workbook
        wb.save(output_path)

        logger.info(f"Excel tracker saved to {output_path}")
        return output_path


# Default tracker instance (lazy loaded)
default_analytics_tracker = None


def get_default_analytics_tracker() -> AnalyticsTracker:
    """Get or create default analytics tracker instance"""
    global default_analytics_tracker
    if default_analytics_tracker is None:
        default_analytics_tracker = AnalyticsTracker()
    return default_analytics_tracker
